import tkinter as tk
from tkinter import PhotoImage, messagebox
from tkinter import ttk
import os
import openpyxl

# Constants
headsize = 40
labelsize = 15
locsize = 20
linewidth = 200

MAX_LAT = 10.6708229
MIN_LAT = 10.6526
MAX_LONG = 77.020833
MIN_LONG = 77.0005

# File path for Excel
excel = r"data/AmbulanceServiceData.xlsx"

# Global variables to store details
lat_val = 0.00
long_val = 0.00
pt_name = ""
pt_gender = ""
relate = ""
opt = ""
pt_id = 0

def writetoexcel(relative, patientname, patientgender, lat, long, option):
    global pt_id  # Use the global pt_id variable
    if not os.path.exists(excel):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Ambulance Requests"
        sheet.append(["Patient ID", "Relation", "Patient Name", "Patient Gender", "Latitude", "Longitude", "Priority"])
        pt_id = 100  # Starting ID if the file does not exist
    else:
        workbook = openpyxl.load_workbook(excel)
        sheet = workbook.active
        max_row = sheet.max_row
        if max_row > 1:
            # Get the last Patient ID
            last_id = sheet.cell(row=max_row, column=1).value
            pt_id = (last_id + 1) if last_id is not None else 100  # Increment or start at 100 if None
        else:
            pt_id = 100  # Default starting ID if there are no previous records

    relationshipstatus = 'Relative' if relative else 'Stranger'
    sheet.append([pt_id, relationshipstatus, patientname, patientgender, lat, long, option])
    workbook.save(excel)

    messagebox.showinfo("Success", f"Data has been successfully saved to the Excel file.\nPatient ID: {pt_id}")


def isrelative():
    if relative.get():
        patientname.grid(row=2, column=0, columnspan=2, padx=70, pady=(10, 0), sticky="w")
        patientnameline.grid(row=3, column=0, columnspan=2, padx=70, pady=(0, 10), sticky="w")
        patientgender.grid(row=4, column=0, columnspan=2, padx=70, pady=(10, 0), sticky="w")
    else:
        patientname.grid_forget()
        patientnameline.grid_forget()
        patientgender.grid_forget()
        patientgender.set('Select gender')

def validate_and_submit(isrelative, patientname, patientgender, lat, long, option):
    try:
        lat_val = float(lat)
        long_val = float(long)
    except ValueError:
        messagebox.showwarning("Input Error", "Latitude and Longitude must be valid numbers.")
        return False

    if isrelative:
        if not patientname or patientname == 'Patient Name':
            messagebox.showwarning("Input Error", "Please enter the patient's name.")
            return False
        if not patientgender or patientgender == 'Select Gender':
            messagebox.showwarning("Input Error", "Please select the patient's gender.")
            return False

    if not MIN_LAT <= lat_val <= MAX_LAT:
        messagebox.showwarning("Input Error", f"Please enter a valid latitude between {MIN_LAT}° N and {MAX_LAT}° N.")
        return False
    if not MIN_LONG <= long_val <= MAX_LONG:
        messagebox.showwarning("Input Error", f"Please enter a valid longitude between {MIN_LONG}° E and {MAX_LONG}° E.")
        return False
    if option == "Select the case" or not option:
        messagebox.showwarning("Input Error", "Please select the priority type.")
        return False

    writetoexcel(isrelative, patientname, patientgender, lat, long, option)
    return True

def submit():
    global lat_val, long_val, pt_name, pt_gender, relate, opt

    patient_name = patientname.get()
    patient_gender = patientgender.get()
    latitude = lat.get()
    longitude = long.get()
    priority = option.get()

    if validate_and_submit(relative.get(), patient_name, patient_gender, latitude, longitude, priority):
        lat_val = float(latitude)
        long_val = float(longitude)
        pt_name = patient_name
        pt_gender = patient_gender
        opt = priority
        relate = 'Relative' if relative.get() else 'Stranger'
        window.destroy()

def get_details():
    return pt_id, relate, pt_name, pt_gender, lat_val, long_val, opt

def reset():
    relative.set(False)
    isrelative()
    patientname.delete(0, tk.END)
    lat.delete(0, tk.END)
    long.delete(0, tk.END)

    set_placeholder(patientname, 'Patient Name')
    set_placeholder(lat, 'Latitude')
    set_placeholder(long, 'Longitude')
    option.set("Select the case")
    patientgender.set("Select Gender")

def add_placeholder(entry, placeholder):
    entry.insert(0, placeholder)
    entry.bind("<FocusIn>", lambda e: remove_placeholder(entry, placeholder))
    entry.bind("<FocusOut>", lambda e: set_placeholder(entry, placeholder))

def remove_placeholder(entry, placeholder):
    if entry.get() == placeholder:
        entry.delete(0, tk.END)
        entry.config(fg='black')

def set_placeholder(entry, placeholder):
    if entry.get() == '':
        entry.insert(0, placeholder)
        entry.config(fg='gray')

window = tk.Tk()
window.title("Ambulance Service System")
window.geometry("925x500+300+200")
window.config(bg="#ffffff")
window.wm_iconbitmap(r"data/icon.ico")

img = PhotoImage(file=r"data/GUI.png")
labelicon = tk.Label(window, image=img, bg='#ffffff')
labelicon.grid(row=0, column=0, rowspan=5, padx=20, pady=200)

frame = tk.Frame(window, width=350, height=350, bg="white")
frame.grid(row=0, column=1, padx=10, pady=10)

headlabel = tk.Label(frame, text="Ambulance Service System", font=("Times Roman", headsize, "bold"), bg="#ffffff", fg="#ce3847")
headlabel.grid(row=0, column=0, columnspan=2, pady=20)

relative = tk.BooleanVar()
relativecheckbox = tk.Checkbutton(frame, text="Patient Details", variable=relative, bg="#ffffff", font=("Arial", 18), command=isrelative)
relativecheckbox.grid(row=1, column=0, padx=10, pady=10, sticky="w")

patientname = tk.Entry(frame, width=25, fg='gray', border=0, bg='white', font=('Helvetica', labelsize))
add_placeholder(patientname, 'Patient Name')
patientnameline = tk.Frame(frame, width=linewidth, height=1, bg='black')

patientgender = ttk.Combobox(frame, width=15, state="readonly", font=('Helvetica', 14), values=['Male','Female','Other'])
patientgender.set('Select gender')

loc = tk.Label(frame, text="Location Details", bg="white", fg="black", font=("Helvetica", locsize))
loc.grid(row=6, column=0, padx=10, pady=10, sticky="w")

lat = tk.Entry(frame, width=35, fg='gray', bg='white', border=0, font=('Helvetica', labelsize))
add_placeholder(lat, 'Latitude')
lat.grid(row=7, column=0, columnspan=2, padx=70, pady=(10, 0), sticky='w')
latline = tk.Frame(frame, width=linewidth, height=1, bg='black')
latline.grid(row=8, column=0, columnspan=2, padx=70, pady=(0, 10), sticky='w')

long = tk.Entry(frame, width=35, fg='gray', bg='white', border=0, font=('Helvetica', labelsize))
add_placeholder(long, 'Longitude')
long.grid(row=9, column=0, columnspan=2, padx=70, pady=(10, 0), sticky='w')
longline = tk.Frame(frame, width=linewidth, height=1, bg='black')
longline.grid(row=10, column=0, columnspan=2, padx=70, pady=(0, 50), sticky='w')

optionlabel = tk.Label(frame, text="Select the Emergency Case", bg='white', fg='black', font=("Helvetica", locsize))
optionlabel.grid(row=11, column=0, columnspan=2, padx=5, pady=5, sticky='w')
option = ttk.Combobox(frame, font=('Helvetica', 14), values=['Heart Attack','Fire Accident','Road Accident', 'Pregnancy', 'Severe Bleeding', 'Stroke', 'Fracture','Respiratory Failure', 'Other'], state='readonly', width=15)
option.set("Select the case")
option.grid(row=12, column=0, columnspan=2, padx=70, pady=(10,0), sticky='w')

submitbutton = tk.Button(frame, text="Submit", bg="#ce3847", fg="white", font=("Helvetica", 18), command=submit)
submitbutton.grid(row=13, column=0, columnspan=2, padx=100, pady=20, sticky='w')

resetbutton = tk.Button(frame, text="Reset", bg="white", fg="#ce3847", font=("Helvetica", 18), command=reset)
resetbutton.grid(row=13, column=1, padx=30, pady=20, sticky='e')

window.mainloop()
